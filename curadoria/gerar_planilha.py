"""
EsquadriaCore — curadoria/gerar_planilha
=========================================
Regenera curadoria/biblioteca_curadoria.xlsx a partir dos JSONs da biblioteca
(dados/geometrias.json + dados/perfil_geometria.json), que são a fonte de
verdade. Rodar após cada homologação.

Uso:  python3 curadoria/gerar_planilha.py
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Pesos dos candidatos comuns Alcoa × Vitral Sul (medição 2026-07-09).
# Alcoa: índice do catálogo (texto nativo). Vitral Sul: OCR das págs raster.
PESOS_CANDIDATOS = {
    "SU-005": (1.108, 1.108),
    "SU-009": (0.908, 0.877),
    "SU-024": (1.024, 0.959),
    "SU-025": (0.973, 0.919),
    "SU-056": (0.539, 0.519),
    "SU-228": (0.688, 0.688),
    "SU-230": (0.954, 0.936),
    "SU-280": (1.006, 0.969),
}
PROVENIENCIA = ("Alcoa: índice do catálogo (texto nativo). Vitral Sul: OCR das "
                "páginas raster (confiança menor). Medição: 2026-07-09.")

HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BASE = Font(name="Arial")
AZUL_INPUT = Font(name="Arial", color="0000FF")
ALERT_FILL = PatternFill("solid", start_color="FFF2CC")


def cabecalho(ws, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font, cell.fill = HDR_FONT, HDR_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def ajustar_larguras(ws, larguras):
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def gerar():
    with open("dados/geometrias.json") as f:
        geos = json.load(f)["geometrias"]
    with open("dados/perfil_geometria.json") as f:
        assocs = json.load(f)["associacoes"]

    homologados = {g["id"].replace("GEO-", "") for g in geos}
    pendentes = [c for c in sorted(PESOS_CANDIDATOS) if c not in homologados]

    wb = Workbook()

    ws = wb.active
    ws.title = "Geometrias"
    cabecalho(ws, ["ID", "Descrição", "Família", "Status", "Versão", "Curado por",
                   "Data", "Largura (mm)", "Altura (mm)", "Esp. trilhos (mm)",
                   "Caixa (mm)", "Queda trilhos (mm)", "Nota"])
    for r, g in enumerate(geos, 2):
        cotas = g.get("cotas_confirmadas_mm", {})
        linha = [g["id"], g["descricao"], g["familia_mercado"], g["status"],
                 g["versao"], g["curado_por"], g["data_curadoria"],
                 cotas.get("largura"), cotas.get("altura"),
                 cotas.get("espacamento_trilhos"), cotas.get("caixa"),
                 cotas.get("queda_trilhos"), g.get("_nota", "")]
        for c, v in enumerate(linha, 1):
            ws.cell(row=r, column=c, value=v).font = BASE
    ajustar_larguras(ws, [14, 42, 12, 18, 8, 12, 12, 12, 11, 15, 10, 16, 70])

    ws = wb.create_sheet("Associações")
    cabecalho(ws, ["Perfil", "Fabricante", "Geometria", "Responsável",
                   "Método de validação", "Data", "Confiança", "Observações"])
    for r, a in enumerate(assocs, 2):
        fab = "Alcoa" if a["perfil_id"].startswith("ALCOA") else "Vitral Sul"
        linha = [a["perfil_id"], fab, a["geometria_padrao_id"],
                 a["responsavel_homologacao"], a["metodo_validacao"],
                 a["data"], a["nivel_de_confianca"], a["observacoes"]]
        for c, v in enumerate(linha, 1):
            ws.cell(row=r, column=c, value=v).font = BASE
    ajustar_larguras(ws, [20, 12, 14, 12, 48, 12, 11, 75])

    ws = wb.create_sheet("Candidatos pendentes")
    cabecalho(ws, ["Código", "Peso Alcoa (kg/m)", "Peso Vitral Sul (kg/m)",
                   "Δ peso", "Sinal", "Proveniência dos pesos"])
    for r, cod in enumerate(pendentes, 2):
        pa, pv = PESOS_CANDIDATOS[cod]
        ws.cell(row=r, column=1, value=cod).font = BASE
        ws.cell(row=r, column=2, value=pa).font = AZUL_INPUT
        ws.cell(row=r, column=3, value=pv).font = AZUL_INPUT
        d = ws.cell(row=r, column=4, value=f"=ABS(B{r}-C{r})/MAX(B{r},C{r})")
        d.font, d.number_format = BASE, "0.0%"
        tol_ref = f"$B${len(pendentes) + 4}"
        s = ws.cell(row=r, column=5,
                    value=f'=IF(D{r}<{tol_ref},"✓ bate","⚠ olhar desenho")')
        s.font = BASE
        ws.cell(row=r, column=6, value=PROVENIENCIA).font = Font(
            name="Arial", size=9, italic=True)

    lin_tol = len(pendentes) + 4
    ws.cell(row=lin_tol, column=1, value="Tolerância Δ peso:").font = Font(
        name="Arial", bold=True)
    tol = ws.cell(row=lin_tol, column=2, value=0.06)
    tol.font, tol.number_format, tol.fill = AZUL_INPUT, "0.0%", ALERT_FILL
    nota = ws.cell(row=lin_tol + 2, column=1,
                   value="Regra de curadoria: peso batendo AUMENTA a confiança; "
                         "divergência é ALERTA para olhar o desenho — nunca veto "
                         "automático. Decisão final é visual e humana (Bruno).")
    nota.font = Font(name="Arial", italic=True)
    ajustar_larguras(ws, [10, 18, 20, 10, 18, 95])

    wb.save("curadoria/biblioteca_curadoria.xlsx")
    print(f"planilha regenerada: {len(geos)} geometrias, {len(assocs)} "
          f"associações, {len(pendentes)} pendentes ({', '.join(pendentes)})")


if __name__ == "__main__":
    gerar()
